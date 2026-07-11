import io
import csv
from datetime import date
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from backend.app.core.database import get_db
from backend.app.core.security import get_current_user
from backend.app.models.models import Order, User, ActivityLog
from backend.app.api.dashboard import apply_filters
from backend.app.services.forecaster import Forecaster
from backend.app.services.ai_service import AIService

router = APIRouter(prefix="/reports", tags=["reports"])

@router.get("/csv")
def export_csv(
    region: Optional[str] = None,
    category: Optional[str] = None,
    sales_rep: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Order)
    query = apply_filters(query, None, None, region, category, sales_rep)
    orders = query.order_by(Order.date.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Order ID", "Date", "Region", "City", "Product", "Category", 
        "Quantity", "Price", "Discount", "Sales", "Profit", 
        "Customer Segment", "Sales Representative", "Is Anomaly"
    ])
    
    # Rows
    for o in orders:
        writer.writerow([
            o.order_id, o.date.strftime("%Y-%m-%d"), o.region, o.city, o.product, o.category,
            o.quantity, o.price, o.discount, o.sales, o.profit,
            o.customer_segment, o.sales_rep, "Yes" if o.is_anomaly else "No"
        ])
        
    output.seek(0)
    
    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action="EXPORT_CSV",
        details=f"Exported CSV report with {len(orders)} orders."
    )
    db.add(log)
    db.commit()
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=insightiq_export_{date.today()}.csv"}
    )

@router.get("/excel")
def export_excel(
    region: Optional[str] = None,
    category: Optional[str] = None,
    sales_rep: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Order)
    query = apply_filters(query, None, None, region, category, sales_rep)
    orders = query.order_by(Order.date.desc()).all()
    
    # Create Workbook
    wb = Workbook()
    
    # Sheet 1: Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"] = "InsightIQ – AI Executive Summary Report"
    ws_summary["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary["A1"].fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo
    ws_summary.row_dimensions[1].height = 40
    
    # Fetch KPI summary
    sales_val = sum(o.sales for o in orders)
    profit_val = sum(o.profit for o in orders)
    orders_count = len(orders)
    aov_val = sales_val / orders_count if orders_count > 0 else 0.0
    anomalies_count = sum(1 for o in orders if o.is_anomaly)
    
    # KPIs Styling & Values
    ws_summary["A3"] = "Performance KPIs"
    ws_summary["A3"].font = Font(name="Segoe UI", size=12, bold=True, color="4F46E5")
    
    headers = ["Metric", "Value", "", "Key Insights / Recommendations", "Action Item"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_summary.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[4].height = 25
    
    kpis = [
        ("Total Sales", f"${sales_val:,.2f}"),
        ("Total Profit", f"${profit_val:,.2f}"),
        ("Profit Margin", f"{(profit_val/sales_val*100) if sales_val > 0 else 0.0:.2f}%"),
        ("Orders Count", f"{orders_count:,}"),
        ("Average Order Value", f"${aov_val:,.2f}"),
        ("Flagged Anomalies", f"{anomalies_count} orders")
    ]
    
    border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for idx, (lbl, val) in enumerate(kpis):
        r = 5 + idx
        c_lbl = ws_summary.cell(row=r, column=1, value=lbl)
        c_val = ws_summary.cell(row=r, column=2, value=val)
        c_lbl.font = Font(name="Segoe UI", size=10)
        c_val.font = Font(name="Segoe UI", size=10, bold=True)
        c_lbl.border = cell_border
        c_val.border = cell_border
        
    # Seed fallback insights directly onto Summary tab
    ai_insights = AIService.generate_insights(db)
    
    ws_summary.merge_cells("D5:D10")
    ws_summary["D5"] = ai_insights.summary
    ws_summary["D5"].font = Font(name="Segoe UI", size=9, italic=True)
    ws_summary["D5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_summary["D5"].border = cell_border
    
    ws_summary["A13"] = "AI Recommendations"
    ws_summary["A13"].font = Font(name="Segoe UI", size=12, bold=True, color="4F46E5")
    
    ws_summary.cell(row=14, column=1, value="Recommendation").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=14, column=1).fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    ws_summary.cell(row=14, column=2, value="Impact / Metric").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=14, column=2).fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    ws_summary.merge_cells("C14:E14")
    ws_summary["C14"] = "Action Details"
    ws_summary["C14"].font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary["C14"].fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    ws_summary.row_dimensions[14].height = 25
    
    for idx, card in enumerate(ai_insights.insights[:4]):
        r = 15 + idx
        ws_summary.cell(row=r, column=1, value=card.title).font = Font(name="Segoe UI", size=9, bold=True)
        ws_summary.cell(row=r, column=2, value=card.metric).font = Font(name="Segoe UI", size=9)
        ws_summary.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws_summary.cell(row=r, column=3, value=f"{card.recommendation} ({card.description})").font = Font(name="Segoe UI", size=9)
        ws_summary.cell(row=r, column=1).border = cell_border
        ws_summary.cell(row=r, column=2).border = cell_border
        ws_summary.cell(row=r, column=3).border = cell_border
        ws_summary.cell(row=r, column=4).border = cell_border
        ws_summary.cell(row=r, column=5).border = cell_border
        
    # Auto-adjust column widths for Executive Summary
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_summary.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
        
    # Sheet 2: Detailed Data
    ws_data = wb.create_sheet(title="Sales Data")
    ws_data.views.sheetView[0].showGridLines = True
    
    headers_data = [
        "Order ID", "Date", "Region", "City", "Product", "Category", 
        "Quantity", "Price", "Discount", "Sales", "Profit", 
        "Customer Segment", "Sales Rep", "Is Anomaly"
    ]
    for col_idx, h in enumerate(headers_data, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
    
    for row_idx, o in enumerate(orders, start=2):
        ws_data.cell(row=row_idx, column=1, value=o.order_id)
        ws_data.cell(row=row_idx, column=2, value=o.date.strftime("%Y-%m-%d"))
        ws_data.cell(row=row_idx, column=3, value=o.region)
        ws_data.cell(row=row_idx, column=4, value=o.city)
        ws_data.cell(row=row_idx, column=5, value=o.product)
        ws_data.cell(row=row_idx, column=6, value=o.category)
        ws_data.cell(row=row_idx, column=7, value=o.quantity)
        ws_data.cell(row=row_idx, column=8, value=o.price)
        ws_data.cell(row=row_idx, column=9, value=o.discount)
        ws_data.cell(row=row_idx, column=10, value=o.sales)
        ws_data.cell(row=row_idx, column=11, value=o.profit)
        ws_data.cell(row=row_idx, column=12, value=o.customer_segment)
        ws_data.cell(row=row_idx, column=13, value=o.sales_rep)
        ws_data.cell(row=row_idx, column=14, value="Yes" if o.is_anomaly else "No")
        
    # Auto-adjust column widths for Data Sheet
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_data.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 30)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action="EXPORT_EXCEL",
        details=f"Exported Excel report with executive summary and {len(orders)} orders."
    )
    db.add(log)
    db.commit()
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=insightiq_report_{date.today()}.xlsx"}
    )

@router.get("/pdf")
def export_pdf(
    region: Optional[str] = None,
    category: Optional[str] = None,
    sales_rep: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Order)
    query = apply_filters(query, None, None, region, category, sales_rep)
    orders = query.order_by(Order.date.desc()).all()
    
    # Overall KPIs
    sales_val = sum(o.sales for o in orders)
    profit_val = sum(o.profit for o in orders)
    orders_count = len(orders)
    aov_val = sales_val / orders_count if orders_count > 0 else 0.0
    anomalies_count = sum(1 for o in orders if o.is_anomaly)
    
    # Setup PDF layout
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Palette Styling
    primary_color = colors.HexColor("#4F46E5") # Indigo
    dark_blue = colors.HexColor("#1E1B4B") # Navy
    text_dark = colors.HexColor("#0F172A") # Slate 900
    
    title_style = ParagraphStyle(
        name="DocTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=20,
        textColor=primary_color,
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        name="DocSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=25
    )
    
    section_title_style = ParagraphStyle(
        name="SecTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=dark_blue,
        spaceBefore=15,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        name="BodyTextCustom",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        textColor=text_dark,
        spaceAfter=10
    )
    
    bullet_style = ParagraphStyle(
        name="BulletCustom",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=text_dark,
        leftIndent=15,
        firstLineIndent=-10,
        spaceAfter=6
    )
    
    elements = []
    
    # Header Title
    elements.append(Paragraph("InsightIQ – AI Decision Intelligence Platform", title_style))
    elements.append(Paragraph(
        f"Executive Performance Report • Generated on {date.today().strftime('%B %d, %Y')} "
        f"• Filters applied: region={region or 'All'}, category={category or 'All'}, rep={sales_rep or 'All'}",
        subtitle_style
    ))
    
    # 1. KPIs Section
    elements.append(Paragraph("Key Performance Indicators", section_title_style))
    
    kpi_data = [
        [Paragraph("<b>Metric</b>", body_style), Paragraph("<b>Performance Value</b>", body_style)],
        [Paragraph("Total Sales", body_style), Paragraph(f"${sales_val:,.2f}", body_style)],
        [Paragraph("Total Profit", body_style), Paragraph(f"${profit_val:,.2f}", body_style)],
        [Paragraph("Profit Margin", body_style), Paragraph(f"{(profit_val/sales_val*100) if sales_val > 0 else 0.0:.2f}%", body_style)],
        [Paragraph("Total Orders", body_style), Paragraph(f"{orders_count:,}", body_style)],
        [Paragraph("Average Order Value", body_style), Paragraph(f"${aov_val:,.2f}", body_style)],
        [Paragraph("Flagged Anomalies", body_style), Paragraph(f"{anomalies_count} orders", body_style)],
    ]
    
    kpi_table = Table(kpi_data, colWidths=[250, 250])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (1,0), colors.HexColor("#F1F5F9")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
    ]))
    
    elements.append(kpi_table)
    elements.append(Spacer(1, 15))
    
    # 2. Forecasting summary
    elements.append(Paragraph("Sales Forecast Summary (Prophet Model)", section_title_style))
    
    # Generate temporary forecast to display projections
    df = pd.DataFrame([{"date": o.date, "sales": o.sales, "profit": o.profit} for o in orders])
    df["date"] = pd.to_datetime(df["date"])
    
    try:
        fc = Forecaster.generate_forecast(df, horizon=30)
        summary = fc["summary"]
        
        fc_data = [
            [Paragraph("<b>Forecast Indicator</b>", body_style), Paragraph("<b>30-Day Projection</b>", body_style)],
            [Paragraph("Projected Total Sales", body_style), Paragraph(f"${summary['projected_sales']:,.2f}", body_style)],
            [Paragraph("Projected Total Profit", body_style), Paragraph(f"${summary['projected_profit']:,.2f}", body_style)],
            [Paragraph("Average Projected Daily Sales", body_style), Paragraph(f"${summary['average_daily_sales']:,.2f}", body_style)],
            [Paragraph("Forecast Trend Direction", body_style), Paragraph(summary["growth_trend_direction"], body_style)],
            [Paragraph("Model Confidence Level", body_style), Paragraph(f"{summary['confidence_level']*100:.0f}%", body_style)],
        ]
        
        fc_table = Table(fc_data, colWidths=[250, 250])
        fc_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor("#F1F5F9")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ]))
        elements.append(fc_table)
    except Exception as e:
        elements.append(Paragraph(f"Forecasting was unable to run due to: {e}", body_style))
        
    elements.append(Spacer(1, 15))
    
    # 3. AI Insights and Recommendations
    elements.append(Paragraph("AI-Generated Recommendations", section_title_style))
    
    ai_res = AIService.generate_insights(db)
    elements.append(Paragraph(f"<b>Executive Brief:</b> {ai_res.summary}", body_style))
    elements.append(Spacer(1, 5))
    
    for idx, card in enumerate(ai_res.insights[:4]):
        rec_text = f"<b>{idx+1}. {card.title} ({card.metric}):</b> {card.recommendation} — <i>{card.description}</i>"
        elements.append(Paragraph(rec_text, bullet_style))
        
    # Build Document
    doc.build(elements)
    buffer.seek(0)
    
    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action="EXPORT_PDF",
        details="Exported PDF Executive Summary report."
    )
    db.add(log)
    db.commit()
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=insightiq_executive_report_{date.today()}.pdf"}
    )
